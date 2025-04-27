import logging
from datetime import datetime
from decimal import Decimal
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, F
from django.db import transaction
from datetime import date
from openpyxl import load_workbook
from .forms import (
    ChartOfAccountsForm,
    IncomeTransactionForm,
    ExpenseTransactionForm,
    TransactionFormSet,
    ImportCOAForm,
)
from apps.sales.models import SaleDetail
from .models import ChartOfAccounts, Transaction
from apps.sales.forms import ReportPeriodForm

from apps.authentication.decorators import (
    admin_or_manager_or_staff_required,
    admin_or_manager_required,
    admin_required,
)


logger = logging.getLogger(__name__)


# =================================== Account List view ===================================
def chart_of_accounts_list_view(request):
    accounts = ChartOfAccounts.objects.all()  # Retrieve all accounts
    accounts_by_type = {}  # Dictionary to group accounts by type

    # Group accounts by their account type
    for account in accounts:
        account_type = account.get_account_type_display()
        if account_type not in accounts_by_type:
            accounts_by_type[account_type] = []
        accounts_by_type[account_type].append(account)

    context = {
        "accounts_by_type": accounts_by_type,
        "table_title": "Chart of Accounts",
    }
    return render(request, "finance/chart_of_accounts_list.html", context)


# =================================== Process and Import Excel data ===================================
@login_required
@admin_required
@transaction.atomic
def import_coa_data(request):
    if request.method == "POST":
        form = ImportCOAForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES.get("excel_file")
            if excel_file and excel_file.name.endswith(".xlsx"):
                try:
                    # Call process_and_import_accounts_data function
                    errors = process_and_import_accounts_data(excel_file)
                    if errors:
                        for error in errors:
                            messages.error(request, error, extra_tags="bg-danger")
                    else:
                        messages.success(
                            request,
                            "Data imported successfully!",
                            extra_tags="bg-success",
                        )
                except Exception as e:
                    messages.error(
                        request, f"Error importing data: {e}", extra_tags="bg-danger"
                    )
                return redirect("finance:chart_of_accounts_list")
            else:
                messages.error(
                    request, "Please upload a valid Excel file.", extra_tags="bg-danger"
                )
    else:
        form = ImportCOAForm()
    return render(
        request,
        "finance/accounts_import.html",
        {"form_name": "Import Accounts - Excel", "form": form},
    )


# Function to import Excel data
@transaction.atomic
def process_and_import_accounts_data(excel_file):
    errors = []
    try:
        wb = load_workbook(excel_file)
        sheet = wb.active

        for row_num, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            account_name = row[0].value
            account_type = row[1].value
            account_number = row[2].value
            description = row[3].value

            # Ensure account_number is treated as a string
            if account_number is None:
                errors.append(f"Missing account number on row {row_num}")
                continue

            # Convert to string, even if it's a number
            account_number = str(account_number)

            if account_name and account_type and account_number:
                try:
                    # Validate account type
                    if (
                        account_type
                        not in dict(ChartOfAccounts.ACCOUNT_TYPE_CHOICES).keys()
                    ):
                        errors.append(
                            f"Invalid account type '{account_type}' on row {row_num}"
                        )
                        continue

                    # Validate that the account number is numeric
                    if not account_number.isdigit():
                        errors.append(
                            f"Account number must be numeric on row {row_num}"
                        )
                        continue

                    # Create the account
                    ChartOfAccounts.objects.create(
                        account_name=account_name,
                        account_type=account_type,
                        account_number=account_number,
                        description=description,
                    )
                except Exception as e:
                    errors.append(f"Error on row {row_num}: {e}")
                    logger.error(f"Error on row {row_num}: {e}")
            else:
                errors.append(f"Missing required fields on row {row_num}")
    except Exception as e:
        errors.append(f"Failed to process the Excel file: {e}")
        logger.error(f"Failed to process the Excel file: {e}")

    return errors


# =================================== Add Account view ===================================
@login_required
@admin_or_manager_required
def add_chart_of_account_view(request):
    form = ChartOfAccountsForm(request.POST or None)

    if form.is_valid():
        form.save()
        messages.success(
            request, "Account added successfully!", extra_tags="bg-success"
        )
        return redirect("finance:add_chart_of_account")

    # Additional context for the template
    context = {
        "form": form,
        "table_title": "Add New Account",
    }

    return render(request, "finance/chart_of_account_add.html", context)


# =================================== Account update view ===================================
@login_required
@admin_or_manager_or_staff_required
@transaction.atomic
def chart_of_account_update_view(request, account_id):
    account = get_object_or_404(ChartOfAccounts, id=account_id)

    if request.method == "POST":
        form = ChartOfAccountsForm(request.POST, instance=account)
        if form.is_valid():
            form.save()
            messages.success(
                request,
                f"Account: {account.account_name} updated successfully!",
                extra_tags="bg-success",
            )
            return redirect("finance:chart_of_accounts_list")
        else:
            messages.error(
                request,
                "There was an error updating the account!",
                extra_tags="bg-danger",
            )
    else:
        form = ChartOfAccountsForm(instance=account)

    context = {"form": form, "account": account, "page_title": "Update Account"}

    return render(request, "finance/chart_of_account_update.html", context)


# =================================== Account delete view ===================================
@login_required
@admin_required
@transaction.atomic
def chart_of_account_delete_view(request, account_id):
    account = get_object_or_404(ChartOfAccounts, id=account_id)

    try:
        account.delete()
        messages.success(
            request,
            f"Account: {account.account_name} deleted successfully!",
            extra_tags="bg-success",
        )
    except Exception as e:
        messages.error(
            request,
            "An error occurred during the deletion process.",
            extra_tags="bg-danger",
        )
        print(f"Error deleting account: {e}")

    return redirect("finance:chart_of_accounts_list")


# =================================== Income transaction creation view ===================================
@login_required
@admin_or_manager_required
@transaction.atomic
def income_transaction_create_view(request):
    if request.method == "POST":
        form = IncomeTransactionForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(
                request,
                "Income transaction posted successfully.",
                extra_tags="bg-success",
            )
            return redirect("finance:income_add")
    else:
        form = IncomeTransactionForm()

    context = {
        "form": form,
        "form_title": "Add New Income Transaction",
    }

    return render(request, "finance/income_add.html", context)


# =================================== expense add view ===================================
@login_required
@admin_or_manager_required
@transaction.atomic
def expense_transaction_create_view(request):
    if request.method == "POST":
        form = ExpenseTransactionForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(
                request,
                "Expense transaction posted successfully.",
                extra_tags="bg-success",
            )
            return redirect("finance:expense_add")
    else:
        form = ExpenseTransactionForm()

    context = {
        "form": form,
        "form_title": "Add New Expense Transaction",
    }

    return render(request, "finance/expense_add.html", context)


# =================================== multi-journal entry view ===================================
@login_required
@admin_or_manager_required  # Assuming this decorator is defined somewhere
@transaction.atomic
def multi_journal_view(request):
    if request.method == "POST":
        formset = TransactionFormSet(request.POST)
        transaction_date = request.POST.get(
            "transaction_date"
        )  # Capture the single transaction date

        if formset.is_valid():
            transactions = formset.save(commit=False)

            # Initialize total debits and credits
            total_debits = 0
            total_credits = 0

            for transaction in transactions:
                # Assign the captured transaction date to each transaction
                transaction.transaction_date = transaction_date

                # Calculate totals based on transaction type
                if transaction.transaction_type == "debit":
                    total_debits += transaction.amount
                elif transaction.transaction_type == "credit":
                    total_credits += transaction.amount

            # Validate that total debits equal total credits
            if total_debits != total_credits:
                messages.error(
                    request,
                    "Total debits must equal total credits.",
                    extra_tags="bg-danger",
                )
                return render(
                    request,
                    "finance/multi_journal_entry_add.html",
                    {
                        "formset": formset,
                        "form_title": "Add New Transactions",
                    },
                )

            # Save all transactions
            for transaction in transactions:
                transaction.save()

            messages.success(
                request, "Transactions posted successfully.", extra_tags="bg-success"
            )
            return redirect("finance:multi_journal")
    else:
        formset = TransactionFormSet(
            queryset=Transaction.objects.none()
        )  # Start with empty forms

    context = {
        "formset": formset,
        "form_title": "Add New Transactions",
    }

    return render(request, "finance/multi_journal_entry_add.html", context)


# =================================== ledger_report ist view ===================================
def get_financial_year_dates():
    """Returns the start and end dates for the current financial year."""
    today = date.today()

    # Check if today is after July 1st (start of the financial year)
    if today.month >= 7:
        start_date = date(today.year, 7, 1)  # July 1st of the current year
        end_date = date(today.year + 1, 6, 30)  # June 30th of the next year
    else:
        start_date = date(today.year - 1, 7, 1)  # July 1st of the previous year
        end_date = date(today.year, 6, 30)  # June 30th of the current year

    return start_date, end_date


@login_required
@admin_or_manager_required
def ledger_report_view(request):
    selected_account_id = request.GET.get("account_id")  # Get selected account ID
    ledger_data = []
    accounts = ChartOfAccounts.objects.all()  # Fetch all accounts for the dropdown
    total_debits = 0
    total_credits = 0

    # Get the start and end dates for the current financial year
    financial_year_start, financial_year_end = get_financial_year_dates()

    # Use query parameters or default to the financial year range
    start_date = request.GET.get("start_date") or financial_year_start
    end_date = request.GET.get("end_date") or financial_year_end

    selected_account = None
    opening_balance = 0

    if selected_account_id:
        selected_account = get_object_or_404(ChartOfAccounts, id=selected_account_id)

        # Get transactions within the selected date range
        ledger_data = Transaction.objects.filter(
            account=selected_account, transaction_date__range=[start_date, end_date]
        ).order_by("transaction_date")

        # Get opening balance by calculating the balance before the start_date
        opening_balance_queryset = Transaction.objects.filter(
            account=selected_account, transaction_date__lt=start_date
        )

        # Calculate the opening balance as the sum of all prior debits and credits
        for transaction in opening_balance_queryset:
            if transaction.transaction_type == "debit":
                opening_balance += transaction.amount
            elif transaction.transaction_type == "credit":
                opening_balance -= transaction.amount

        # Calculate debits, credits, and running balance
        running_balance = opening_balance
        for transaction in ledger_data:
            if transaction.transaction_type == "debit":
                transaction.debit = transaction.amount
                transaction.credit = 0
                total_debits += transaction.amount
            elif transaction.transaction_type == "credit":
                transaction.debit = 0
                transaction.credit = transaction.amount
                total_credits += transaction.amount
            else:
                transaction.debit = 0
                transaction.credit = 0

            # Update running balance
            running_balance += transaction.debit - transaction.credit
            transaction.running_balance = running_balance

    return render(
        request,
        "finance/ledger_report.html",
        {
            "ledger_data": ledger_data,
            "accounts": accounts,
            "selected_account": selected_account,
            "selected_account_id": selected_account_id,
            "start_date": start_date,
            "end_date": end_date,
            "total_debits": total_debits,
            "total_credits": total_credits,
            "opening_balance": opening_balance,  # Pass opening balance to template
        },
    )


# =================================== profit_and_loss_view ===================================
@login_required
@admin_or_manager_required
def profit_and_loss_view(request):
    # Default date range values
    start_date = None
    end_date = None

    # Initialize the profit and loss sections with dynamic sections
    profit_and_loss = {
        "Income": [],
        "Expenses": [],
        "Summary": [
            {"label": "Gross Profit", "value": 0.0},
            {"label": "Net Profit", "value": 0.0},
        ],
    }

    # Initialize the form
    form = ReportPeriodForm(request.GET or None)

    if form.is_valid():
        # Extract the start and end dates from the form
        start_date = form.cleaned_data["start_date"]
        end_date = form.cleaned_data["end_date"]

        # Filter data based on the date range
        # Step 1: Query the related SaleDetails for the desired Sale objects
        sales_details = SaleDetail.objects.filter(
            sale__trans_date__range=(start_date, end_date)
        )

        expense_transactions = Transaction.objects.filter(
            transaction_date__range=(start_date, end_date),
            account__account_type="expense",
        )
        other_income_transactions = Transaction.objects.filter(
            transaction_date__range=(start_date, end_date),
            account__account_type="revenue",
        ).exclude(account__account_name="Sales Revenue")

        # Step 2: Aggregate total revenue (Sum of quantity * price)
        total_revenue = sales_details.aggregate(
            total_revenue=Sum(F("quantity") * F("price"))
        )

        # Step 3: Extract the total revenue value from the dictionary
        # Ensure it is a Decimal (if None, set to Decimal("0.00"))
        revenue_value = Decimal(total_revenue["total_revenue"] or "0.00")

        # Step 4: Calculate Other Income (Sum of amounts for income transactions)
        other_income = other_income_transactions.aggregate(total_income=Sum("amount"))[
            "total_income"
        ] or Decimal(
            "0.00"
        )  # Default to Decimal(0.00) if None or no data

        # Step 5: Calculate Total Income (Revenue + Other Income)
        total_income = revenue_value + other_income

        # Dynamic Operating Expenses Calculation (Aggregate by Account Name or Account Type)
        expense_categories = (
            expense_transactions.values(
                "account__account_name", "account__account_number"
            )
            .annotate(total_expense=Sum("amount"))
            .order_by("account__account_number")
        )

        operating_expenses = {}
        for category in expense_categories:
            account_name = category["account__account_name"]
            account_number = category["account__account_number"]
            operating_expenses[f"{account_number} - {account_name}"] = category[
                "total_expense"
            ]

        total_expenses = sum(operating_expenses.values())

        # Calculate COGS (Cost of Goods Sold)
        cogs = (
            SaleDetail.objects.filter(sale__trans_date__range=[start_date, end_date])
            .annotate(cogs=F("product_volume__volume__cost") * F("quantity"))
            .aggregate(total_cogs=Sum("cogs"))
        )["total_cogs"] or 0

        # Calculate Gross Profit (Revenue - COGS)
        gross_profit = revenue_value - cogs

        # Calculate Net Profit (Gross Profit + Other Income - Operating Expenses)
        net_profit = gross_profit + other_income - total_expenses

        # Update the profit_and_loss dictionary with dynamic values
        # Income Section
        profit_and_loss["Income"].append(
            {"label": "Sales Revenue", "value": revenue_value}
        )

        # Split Other Income into categories (if applicable)
        other_income_categories = (
            other_income_transactions.values(
                "account__account_name", "account__account_number"
            )
            .annotate(total_other_income=Sum("amount"))
            .order_by("account__account_number")
        )
        for income_category in other_income_categories:
            account_name = income_category["account__account_name"]
            account_number = income_category["account__account_number"]
            profit_and_loss["Income"].append(
                {
                    "label": f"{account_number} - {account_name}",
                    "value": income_category["total_other_income"],
                }
            )

        profit_and_loss["Income"].append(
            {"label": "Total Other Income", "value": other_income}
        )
        profit_and_loss["Income"].append(
            {"label": "Total Income", "value": total_income}
        )

        # Expenses Section (Dynamic categories from operating_expenses)
        for category, value in operating_expenses.items():
            profit_and_loss["Expenses"].append({"label": category, "value": value})

        # Add the total expenses
        profit_and_loss["Expenses"].append(
            {"label": "Total Expenses", "value": total_expenses}
        )

        # Add Cost of Goods Sold (COGS)
        profit_and_loss["Expenses"].append(
            {"label": "Cost of Goods Sold (COGS)", "value": cogs}
        )

        # Summary Section
        profit_and_loss["Summary"][0]["value"] = gross_profit
        profit_and_loss["Summary"][1]["value"] = net_profit

    # Pass data to the template
    context = {
        "form": form,
        "profit_and_loss": profit_and_loss,
        "start_date": start_date,
        "end_date": end_date,
        "table_title": "Profit and Loss Statement",
    }
    return render(request, "finance/profit_and_loss.html", context)


# =================================== Balnce Sheet View ===================================
# @login_required
# @admin_or_manager_required
# def balance_sheet_view(request):
#     # Set the date range (start_date, end_date) based on user input or defaults
#     start_date = request.GET.get("start_date", datetime.today().strftime("%Y-%m-%d"))
#     end_date = request.GET.get("end_date", datetime.today().strftime("%Y-%m-%d"))

#     # Convert string dates to date objects
#     start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
#     end_date = datetime.strptime(end_date, "%Y-%m-%d").date()

#     # Fetch transactions for the selected date range
#     transactions = Transaction.objects.filter(
#         transaction_date__range=[start_date, end_date]
#     )

#     # Initialize the balances
#     assets = 0
#     liabilities = 0
#     equity = 0
#     revenue = 0
#     expenses = 0

#     # Calculate the balances based on account type
#     for account in ChartOfAccounts.objects.all():
#         account_transactions = transactions.filter(account=account)

#         # Calculate the net balance for the account
#         debit_total = (
#             account_transactions.filter(transaction_type="debit").aggregate(
#                 Sum("amount")
#             )["amount__sum"]
#             or 0
#         )
#         credit_total = (
#             account_transactions.filter(transaction_type="credit").aggregate(
#                 Sum("amount")
#             )["amount__sum"]
#             or 0
#         )
#         net_balance = debit_total - credit_total

#         total_amount = account_transactions.aggregate(Sum("amount"))["amount__sum"] or 0

#         if account.account_type == "asset":
#             assets += net_balance
#         elif account.account_type == "liability":
#             liabilities += net_balance
#         elif account.account_type == "equity":
#             equity += total_amount
#         elif account.account_type == "revenue":
#             revenue += total_amount
#         elif account.account_type == "expense":
#             expenses += total_amount

#     # Fetch SaleDetails within the date range
#     sales_details = SaleDetail.objects.filter(
#         sale__trans_date__range=(start_date, end_date)
#     )

#     # Calculate Revenue: Sum of (selling price * quantity sold)
#     sales_revenue = (
#         sales_details.annotate(selling_price=F("product_volume__volume__price")).aggregate(
#             total_revenue=Sum(F("selling_price") * F("quantity"))
#         )["total_revenue"]
#         or 0
#     )

#     # Calculate Cost of Goods Sold (COGS)
#     cogs = (
#         sales_details.annotate(cost_price=F("product_volume__volume__cost")).aggregate(
#             total_cogs=Sum(F("cost_price") * F("quantity"))
#         )["total_cogs"]
#         or 0
#     )

#     # Calculate Gross Profit (Sales Revenue - COGS)
#     gross_profit = sales_revenue - cogs

#     # Calculate Net Profit (Gross Profit + Other Income - Operating Expenses)
#     net_income = gross_profit + revenue - expenses

#     # Calculate Retained Earnings: Retained Earnings = Net Income + Previous Retained Earnings
#     retained_earnings = equity + net_income

#     # Ensure assets = liabilities + equity
#     liabilities = assets - equity

#     # Return the result to the template
#     context = {
#         "start_date": start_date,
#         "end_date": end_date,
#         "assets": assets,
#         "liabilities": liabilities,
#         "equity": equity,
#         "retained_earnings": retained_earnings,
#         "net_income": net_income,
#         "table_title": "Statement of Financial Position",
#     }

#     return render(request, "finance/balance_sheet.html", context)


@login_required
@admin_or_manager_required
def balance_sheet_view(request):
    # Set the date range (start_date, end_date) based on user input or defaults
    start_date = request.GET.get("start_date", datetime.today().strftime("%Y-%m-%d"))
    end_date = request.GET.get("end_date", datetime.today().strftime("%Y-%m-%d"))

    # Convert string dates to date objects
    start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
    end_date = datetime.strptime(end_date, "%Y-%m-%d").date()

    # Fetch transactions for the selected date range
    transactions = Transaction.objects.filter(
        transaction_date__range=[start_date, end_date]
    )

    # Initialize the balances
    assets = 0
    liabilities = 0
    equity = 0
    revenue = 0
    expenses = 0

    # Calculate the balances based on account type
    for account in ChartOfAccounts.objects.all():
        account_transactions = transactions.filter(account=account)

        # Calculate the net balance for the account
        debit_total = (
            account_transactions.filter(transaction_type="debit").aggregate(
                Sum("amount")
            )["amount__sum"]
            or 0
        )
        credit_total = (
            account_transactions.filter(transaction_type="credit").aggregate(
                Sum("amount")
            )["amount__sum"]
            or 0
        )
        net_balance = debit_total - credit_total

        total_amount = account_transactions.aggregate(Sum("amount"))["amount__sum"] or 0

        if account.account_type == "asset":
            assets += net_balance
        elif account.account_type == "liability":
            liabilities += net_balance
        elif account.account_type == "equity":
            equity += total_amount
        elif account.account_type == "revenue":
            revenue += total_amount
        elif account.account_type == "expense":
            expenses += total_amount

    # Fetch SaleDetails within the date range
    sales_details = SaleDetail.objects.filter(
        sale__trans_date__range=(start_date, end_date)
    )

    # Initialize sales revenue
    sales_revenue = 0

    # Loop through each SaleDetail and apply the discount logic
    for sale_detail in sales_details:
        # Get the corresponding ProductVolume instance
        product_volume = sale_detail.product_volume
        discounted_price = product_volume.get_discounted_price()

        # Accumulate revenue (discounted price * quantity)
        sales_revenue += discounted_price * sale_detail.quantity

    # Calculate Cost of Goods Sold (COGS) with discounts considered for prices
    cogs = (
        sales_details.annotate(cost_price=F("product_volume__volume__cost")).aggregate(
            total_cogs=Sum(F("cost_price") * F("quantity"))
        )["total_cogs"]
        or 0
    )

    # Calculate Gross Profit (Sales Revenue - COGS)
    gross_profit = sales_revenue - cogs

    # Calculate Net Profit (Gross Profit + Other Income - Operating Expenses)
    net_income = gross_profit + revenue - expenses

    # Calculate Retained Earnings: Retained Earnings = Net Income + Previous Retained Earnings
    retained_earnings = equity + net_income

    # Ensure assets = liabilities + equity
    liabilities = assets - equity

    # Return the result to the template
    context = {
        "start_date": start_date,
        "end_date": end_date,
        "assets": assets,
        "liabilities": liabilities,
        "equity": equity,
        "retained_earnings": retained_earnings,
        "net_income": net_income,
        "table_title": "Statement of Financial Position",
    }

    return render(request, "finance/balance_sheet.html", context)
